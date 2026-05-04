"""生成 RAG 评估 Excel Review 文件"""

import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]

ID_PREFIX_MAP = {
    "cpu": "CPU 告警",
    "mem": "内存告警",
    "disk": "磁盘告警",
    "slow": "慢响应",
    "svc": "服务不可用",
}
DIFFICULTY_MAP = {"easy": "简单", "medium": "中等", "hard": "困难"}

# ── load data ──────────────────────────────────────────────────
results_file = REPO_ROOT / "docs" / "rag_eval_results_normal.json"
dataset_file = REPO_ROOT / "docs" / "rag_eval_dataset.json"

with open(results_file, "r", encoding="utf-8") as f:
    payload = json.load(f)
with open(dataset_file, "r", encoding="utf-8") as f:
    dataset = json.load(f)

summary = payload["summary"]
results = payload["results"]
dataset_items = {it["id"]: it for it in dataset["items"]}

# ── styles ─────────────────────────────────────────────────────
header_font_white = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
cell_font = Font(name="微软雅黑", size=10)
title_font = Font(name="微软雅黑", bold=True, size=14)
metric_label_font = Font(name="微软雅黑", bold=True, size=11)
metric_value_font = Font(name="微软雅黑", size=11)
header_font = Font(name="微软雅黑", bold=True, size=11)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center")
hit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_cell(ws, row, col, value=None, font=None, align=None, fill=None, fmt=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = font or cell_font
    cell.alignment = align or wrap_align
    cell.border = thin_border
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# Sheet 1 — 概览
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "概览"

ws1.merge_cells("A1:B1")
style_cell(ws1, 1, 1, "SuperBizAgent RAG 评测 - 正常模式", font=title_font, align=center_align)
ws1.row_dimensions[1].height = 30

meta = [
    ("数据集", payload["dataset"]),
    ("Base URL", payload["base_url"]),
    ("检索 Top-K", payload["retrieval_k"]),
    ("题目数", summary["total"]),
    ("评测模式", summary["mode"]),
    ("评测日期", date.today().strftime("%Y-%m-%d")),
]
row = 3
for label, val in meta:
    style_cell(ws1, row, 1, label, font=metric_label_font)
    style_cell(ws1, row, 2, val, font=metric_value_font)
    row += 1

row += 1
metrics = [
    ("Hit@3", f"{summary['hit_at_3']:.2%}"),
    ("Hit@5", f"{summary['hit_at_5']:.2%}"),
    ("MRR", f"{summary['mrr']:.4f}"),
    ("平均要点命中率", f"{summary['avg_point_hit_rate']:.2%}"),
    ("平均响应时间", f"{summary['avg_latency_seconds']:.2f}s"),
    ("P95 响应时间", f"{summary['p95_latency_seconds']:.2f}s"),
    ("混合检索", "已启用 (dense + BM25 二路召回 + RRF 合并)"),
]
for label, val in metrics:
    style_cell(ws1, row, 1, label, font=metric_label_font)
    style_cell(ws1, row, 2, val, font=metric_value_font)
    row += 1

# category breakdown
row += 1
style_cell(ws1, row, 1, "分类", font=header_font)
style_cell(ws1, row, 2, "数量", font=header_font)
style_header_row(ws1, row, 2)
row += 1

cat_count = {}
for r in results:
    prefix = r["id"].split("_")[0]
    cat_name = ID_PREFIX_MAP.get(prefix, prefix)
    cat_count[cat_name] = cat_count.get(cat_name, 0) + 1

for cat in ["CPU 告警", "磁盘告警", "内存告警", "慢响应", "服务不可用"]:
    style_cell(ws1, row, 1, cat, font=cell_font)
    style_cell(ws1, row, 2, cat_count.get(cat, 0), font=cell_font)
    row += 1

# recall path breakdown
row += 1
path_labels = {
    "both": "dense + keyword (两路命中)",
    "dense": "仅 Dense 向量",
    "keyword": "仅 BM25 关键词",
}
style_cell(ws1, row, 1, "召回路径分布", font=header_font)
style_cell(ws1, row, 2, "文档数", font=header_font)
style_header_row(ws1, row, 2)
row += 1
for path_key in ["both", "dense", "keyword"]:
    cnt = summary.get("recall_path_distribution", {}).get(path_key, 0)
    style_cell(ws1, row, 1, path_labels.get(path_key, path_key), font=cell_font)
    style_cell(ws1, row, 2, cnt, font=cell_font)
    row += 1

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 48

# ═══════════════════════════════════════════════════════════════
# Sheet 2 — 逐题详情
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("逐题详情")

headers2 = [
    "序号", "ID", "分类", "难度", "问题",
    "检索文档数", "Hit@3", "Hit@5", "Reciprocal Rank", "要点命中率",
    "响应时间(s)", "Gold 文档", "Top-3 召回",
    "Gold Points", "Matched Points", "模型回答",
    "召回模式", "召回路径统计",
]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(headers2))
ws2.row_dimensions[1].height = 25

for i, r in enumerate(results, 1):
    row = i + 1
    item = dataset_items.get(r["id"], {})
    prefix = r["id"].split("_")[0]
    cat_name = ID_PREFIX_MAP.get(prefix, prefix)

    # Top-3 retrieved docs
    top3_docs = sorted(r.get("retrieved_docs", []), key=lambda d: d.get("rank", 99))[:3]
    top3_lines = []
    for doc in top3_docs:
        fn = doc.get("file_name", "")
        hd = " | ".join(doc.get("headers", []))
        rp = doc.get("recall_path", "")
        top3_lines.append(f"#{doc.get('rank','')} {fn} / {hd} / path={rp}")
    top3_text = "\n".join(top3_lines)

    gold_pts = "\n".join(f"- {p}" for p in r.get("gold_answer_points", []))
    matched_pts = "\n".join(f"- {p}" for p in r.get("matched_points", []))

    row_data = [
        i,
        r["id"],
        cat_name,
        DIFFICULTY_MAP.get(item.get("difficulty", ""), item.get("difficulty", "")),
        r["question"],
        len(r.get("retrieved_docs", [])),
        1 if r.get("hit_at_3") else 0,
        1 if r.get("hit_at_5") else 0,
        r.get("reciprocal_rank", 0),
        r.get("point_hit_rate", 0),
        r.get("latency_seconds", 0),
        "\n".join(r.get("gold_docs", [])),
        top3_text,
        gold_pts,
        matched_pts,
        r.get("answer", ""),
        r.get("recall_mode", ""),
        json.dumps(r.get("recall_path_stats", {}), ensure_ascii=False),
    ]

    for c, val in enumerate(row_data, 1):
        cell = style_cell(ws2, row, c, val)
        if c in (7, 8):  # Hit@3, Hit@5
            cell.alignment = center_align
            cell.fill = hit_fill if val == 1 else miss_fill
        elif c in (6, 9, 10, 11):
            cell.alignment = center_align
        elif c == 10:
            cell.number_format = "0.00%"
        elif c in (9, 11):
            cell.number_format = "0.00"

    ws2.row_dimensions[row].height = max(80, 15 * len(top3_lines))

widths2 = [5, 10, 10, 6, 45, 8, 7, 7, 10, 10, 10, 25, 40, 35, 35, 50, 10, 15]
for c, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# Sheet 3 — 召回明细
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("召回明细")

headers3 = ["ID", "问题", "rank", "file_name", "headers", "queries", "recall_path", "preview"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(headers3))
ws3.row_dimensions[1].height = 25

row = 2
for r in results:
    for doc in r.get("retrieved_docs", []):
        queries = "\n".join(doc.get("queries", []))
        row_data = [
            r["id"],
            r["question"],
            doc.get("rank", ""),
            doc.get("file_name", ""),
            " / ".join(doc.get("headers", [])),
            queries,
            doc.get("recall_path", ""),
            doc.get("preview", ""),
        ]
        for c, val in enumerate(row_data, 1):
            style_cell(ws3, row, c, val)
        ws3.row_dimensions[row].height = 50
        row += 1

widths3 = [10, 35, 5, 20, 25, 30, 12, 60]
for c, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

ws3.freeze_panes = "A2"

# ── save ───────────────────────────────────────────────────────
out_path = REPO_ROOT / "docs" / f"rag_eval_results_{date.today().strftime('%Y-%m-%d')}_normal_review.xlsx"
wb.save(str(out_path))
print(f"OK: {out_path}")
print(f"Sheets: {wb.sheetnames}, rows: {ws1.max_row}/{ws2.max_row}/{ws3.max_row}")
